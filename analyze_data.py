import openpyxl
import os
from collections import Counter

folder = 'Candidate Data'

# Extract all company names from interviews
all_companies = []
all_candidates = []
interviews = 0

for f in os.listdir(folder):
    if 'Interview Sheet' in f and f.endswith('.xlsx'):
        try:
            wb = openpyxl.load_workbook(os.path.join(folder, f), data_only=True)
            for sheet_name in wb.sheetnames:
                if 'Dashboard' not in sheet_name:
                    ws = wb[sheet_name]
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if row and any(cell for cell in row):
                            interviews += 1
                            # Candidate
                            if len(row) > 1 and row[1] and isinstance(row[1], str):
                                all_candidates.append(row[1].strip())
                            # Company - column 6
                            if len(row) > 6 and row[6] and isinstance(row[6], str):
                                all_companies.append(row[6].strip())
        except Exception as e:
            pass

company_counts = Counter(all_companies)
candidate_counts = Counter(all_candidates)

# Write results
with open('data_summary.txt', 'w') as f:
    f.write(f"TOTAL INTERVIEWS LOGGED: {interviews}\n\n")
    f.write(f"UNIQUE CANDIDATES: {len(candidate_counts)}\n\n")
    f.write(f"UNIQUE COMPANIES: {len(company_counts)}\n\n")
    f.write("TOP 50 COMPANIES BY INTERVIEW COUNT:\n")
    for comp, count in sorted(company_counts.items(), key=lambda x: -x[1])[:50]:
        f.write(f"  {comp}: {count}\n")
    f.write("\n\nCANDIDATES:\n")
    for cand, count in sorted(candidate_counts.items(), key=lambda x: -x[1])[:50]:
        f.write(f"  {cand}: {count} interviews\n")

print("Done - check data_summary.txt")
